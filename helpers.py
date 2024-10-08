import pandas as pd
import os
import re
import logging
import subprocess
from datetime import datetime
import csv
from bot import *
from openpyxl import Workbook
import xlrd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenPyXLImage
from telebot.types import Message
import io
from PIL import Image
from io import BytesIO
import openpyxl
import subprocess

def convert_xls_to_xlsx(xls_file):
    try:
        # Menggunakan LibreOffice untuk mengonversi file .xls ke .xlsx
        result = subprocess.run(['libreoffice', '--headless', '--convert-to', 'xlsx', xls_file], check=True)
        xlsx_file = xls_file.replace('.xls', '.xlsx')
        if os.path.exists(xlsx_file):
            return xlsx_file
        else:
            logging.error("File .xlsx tidak ditemukan setelah konversi.")
            return None
    except subprocess.CalledProcessError as e:
        logging.error(f"Error during conversion: {e}")
        return None

def extract_images_from_excel(excel_file):
    """
    Ekstrak gambar dari file Excel dan mengembalikannya dalam format yang didukung oleh Telegram.
    """
    images = []
    try:
        wb = load_workbook(excel_file)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            if hasattr(ws, '_images'):
                for image in ws._images:
                    image_bytes = image._data()  # Mengambil data gambar sebagai bytes
                    img_stream = BytesIO(image_bytes)
                    
                    # Deteksi format gambar menggunakan PIL
                    img = Image.open(img_stream)
                    
                    # Simpan gambar sementara ke BytesIO untuk dikirim melalui bot
                    img_output = BytesIO()
                    img_format = img.format if img.format else 'PNG'  # Default ke PNG jika format tidak diketahui
                    img.save(img_output, format=img_format)
                    img_output.seek(0)  # Kembali ke awal stream
                    
                    images.append(img_output)
        wb.close()
    except Exception as e:
        logging.error("Error extracting images from Excel file: ", exc_info=True)
    
    return images

def count_vcf_contacts(filename):
    try:
        with open(filename, 'r', encoding='utf-8') as file:
            contacts = 0
            for line in file:
                # Each contact starts with "BEGIN:VCARD"
                if line.startswith("BEGIN:VCARD"):
                    contacts += 1
            return contacts
    except FileNotFoundError:
        print(f"File not found: {filename}")
        return 0
    except Exception as e:
        print(f"An error occurred: {e}")
        return 0


def convert2(data):
    try:
        logging.info("Memulai proses konversi")

        # Baca data dari file .txt
        with open(data['filename'], 'r') as file:
            contacts = file.readlines()

        logging.info(f"Kontak yang dibaca dari file: {contacts}")

        # Hilangkan baris kosong dan whitespace
        contacts = [contact.strip() for contact in contacts if contact.strip()]
        logging.info(f"Kontak setelah dihapus whitespace: {contacts}")

        contacts_per_file = data['totalc']  # Jumlah kontak per file (input pengguna)
        total_files = data['totalf']  # Total file yang ingin dihasilkan (input pengguna)
        cname = data['cname']  # Nama kontak yang diinput pengguna

        logging.info(f"Jumlah kontak per file: {contacts_per_file}, Total file: {total_files}, Nama kontak: {cname}")

        # Hitung total kontak yang diperlukan dan sesuaikan jumlah file jika perlu
        total_contacts = len(contacts)
        required_files = (total_contacts + contacts_per_file - 1) // contacts_per_file
        if total_files < required_files:
            logging.info(f"Jumlah file yang diinputkan kurang dari yang diperlukan. Menyesuaikan jumlah file menjadi {required_files}.")
            total_files = required_files

        # Ambil parameter pergantian nama file jika diatur
        change_every = data.get('change_every', None)
        change_limit = data.get('change_limit', None)
        new_names = data.get('new_names', [])

        logging.info(f"change_every: {change_every}, change_limit: {change_limit}, new_names: {new_names}")

        files_created = []
        file_count = 1
        contact_index = 0  # Melacak indeks kontak
        current_name_idx = 0  # Mengatur indeks nama file
        current_file_num = 1  # Menyimpan nomor file untuk setiap nama

        # Mulai membagi file
        for file_idx in range(total_files):
            # Pergantian nama file setiap beberapa file
            if change_every and file_count > change_every and current_name_idx < len(new_names):
                current_name_idx += 1
                current_file_num = 1  # Reset nomor file ketika nama file berubah
                file_count = 1  # Reset hitungan file

            # Tentukan nama file dengan spasi, bukan underscore
            if current_name_idx < len(new_names):
                output_file_name = f"{new_names[current_name_idx]} {current_file_num}.vcf"
            else:
                output_file_name = f"{data['name']} {current_file_num}.vcf"

            logging.info(f"Membuat file: {output_file_name}")

            with open(output_file_name, 'w') as out_file:
                # Tuliskan kontak dalam format VCF
                for i in range(contacts_per_file):
                    if contact_index >= len(contacts):
                        break  # Jika tidak ada lagi kontak untuk ditulis

                    contact_name = f"{cname} {i + 1}"

                    logging.info(f"Menambahkan kontak: {contact_name} dengan nomor {contacts[contact_index]}")

                    # Tuliskan kontak dalam format VCF
                    out_file.write(
                        f"BEGIN:VCARD\nVERSION:3.0\nFN:{contact_name}\nTEL;TYPE=CELL:{contacts[contact_index]}\nEND:VCARD\n"
                    )
                    contact_index += 1  # Naikkan indeks kontak

            files_created.append(output_file_name)
            file_count += 1
            current_file_num += 1  # Naikkan nomor file untuk nama yang sama

        logging.info(f"File yang dihasilkan: {files_created}")
        return files_created

    except Exception as e:
        logging.error(f"Error during conversion: {e}", exc_info=True)
        return []


def rearrange_to_one_column(input_file, output_file):
    try:
        # Baca isi file dan simpan setiap baris sebagai list yang mewakili kolom
        with open(input_file, 'r') as file:
            lines = [line.strip().split() for line in file.readlines()]
        
        # Pastikan file tidak kosong
        if not lines:
            print("File kosong.")
            return

        # Tentukan jumlah baris dan kolom
        rows = len(lines)
        columns = max(len(line) for line in lines)  # Pastikan jumlah kolom sesuai dengan kolom terbanyak

        # Buat list untuk menyimpan hasil pengurutan kolom ke satu kolom
        rearranged = []

        # Urutkan data secara siklis dari atas ke kiri lalu kebawah
        for col in range(columns):
            for row in range(rows):
                if col < len(lines[row]):
                    rearranged.append(lines[row][col])

        # Tulis hasil pengurutan ke file output, setiap item pada baris baru
        with open(output_file, 'w') as file:
            for item in rearranged:
                file.write(item + '\n')
        
        print(f"File berhasil diproses dan disimpan ke {output_file}")
    
    except Exception as e:
        print(f"Terjadi kesalahan: {e}")

# Contoh penggunaan saat script dijalankan langsung
if __name__ == "__main__":
    input_filename = 'input.txt'
    output_filename = 'output.txt'
    rearrange_to_one_column(input_filename, output_filename)

def check_user(wl, user_id):
  if user_id == owner:
    return True

  if str(user_id) not in wl.keys():
    return False
  else:
    now = datetime.now(wib)
    exp_time = datetime.strptime(wl[str(user_id)], datetime_format).replace(tzinfo=wib)

    if now > exp_time:
      return False
    else:
      return True

def convert(data):
    numbers = check_number(data['filename'])
    split_number = split(numbers, data['totalc'])

    countc = 0
    countf = 0
    vcf_files = []
    sisa = []
    
    for numbers in split_number:
        vcard_entries = []
        for number in numbers:
            countc += 1
            vcard_entry = f"BEGIN:VCARD\nVERSION:3.0\nFN:{data['cname']}-{countc}\nTEL;TYPE=CELL:+{number}\nEND:VCARD"
            vcard_entries.append(vcard_entry)

        countf += 1
        if countf > data['totalf']:
            sisa.extend(numbers)  # Use extend instead of append to flatten the list
        else:
            vcf_name = f"files/{data['name']}_{countf}.vcf"
            vcf_files.append(vcf_name)
            
            with open(vcf_name, 'w', encoding='utf-8') as vcard_file:
                vcard_file.write("\n".join(vcard_entries) + "\n")
    
    if sisa:
        file_txt = "files/sisa.txt"
        vcf_files.append(file_txt)
        
        with open(file_txt, 'w', encoding='utf-8') as file:
            file.write("\n".join(sisa) + "\n")
    
    return vcf_files

def convert_vcf(data):
    data['filename'] = convert_xlsx_to_txt(data)
    numbers = check_number(data['filename'])
    split_number = split(numbers, data['totalc'])

    countc = 0
    countf = 0
    vcf_files = []
    
    for numbers in split_number:
        vcard_entries = []
        for number in numbers:
            countc += 1
            vcard_entry = f"BEGIN:VCARD\nVERSION:3.0\nFN:{data['cname']}-{countc}\nTEL;TYPE=CELL:+{number}\nEND:VCARD"
            vcard_entries.append(vcard_entry)

        countf += 1
        vcf_name = f"files/{data['name']}_{countf}.vcf"
        vcf_files.append(vcf_name)
        
        with open(vcf_name, 'w', encoding='utf-8') as vcard_file:
            vcard_file.write("\n".join(vcard_entries) + "\n")
        
        if countf == data['totalf']:
            break

    return vcf_files

def convert_xlsx_to_txt(data):
    df = pd.read_excel(data['filename'])
    file_name = f"files/{data['name']}.txt"
    df.to_csv(file_name, index=False, sep='\t')

    return file_name

def check_number(filename):
    """Baca file dan ambil nomor telepon dari setiap baris."""
    numbers = []
    with open(filename, 'r', encoding='utf-8') as file:
        for line in file:
            line = line.strip()
            if line.isdigit():
                numbers.append(line)
    return numbers
        
def pecah_txt(data):
  numbers = check_number(data['filename'])
  split_number = split(numbers, data['totaln'])
  countf = 0
  files = []

  for numbers in split_number:
    countf+=1
    txt_name = f"files/{data['name']}_{countf}.txt"
    files.append(txt_name)

    with open(txt_name, 'w', encoding='utf-8') as file:
      for number in numbers:
        file.write(number + "\n")

    if countf == data['totalf']:
      break
  
  return files

def pecah_vcf(data):
    with open(data['filename'], 'r', encoding='utf-8') as file:
        lines = file.readlines()

    contacts = []
    current_contact = []

    for line in lines:
        if not line.strip():
            continue

        current_contact.append(line)
        if line.strip() == 'END:VCARD':
            contacts.append(current_contact)
            current_contact = []

    split_contact = split(contacts, data['totalc'])
    countf = 0
    files = []

    for contacts in split_contact:
        countf += 1
        file_name = f"files/{data['name']}_{countf}.vcf"
        files.append(file_name)
        
        with open(file_name, 'w', encoding='utf-8') as file:
            file.write("".join("".join(contact) for contact in contacts))

        if countf == data['totalf']:
            break

    return files

def convert_vcf_to_txt(data):
    vcf_file = data.get('filename')
    txt_file = f"files/{data.get('name')}.txt"

    if not vcf_file or not os.path.isfile(vcf_file):
        raise FileNotFoundError(f"File VCF tidak ditemukan: {vcf_file}")

    try:
        with open(vcf_file, 'r', encoding='utf-8') as vcf_file_content:
            vcf_data = vcf_file_content.read()

        # Memproses data VCF untuk mengekstrak nama dan nomor telepon
        lines = vcf_data.split('END:VCARD')
        with open(txt_file, 'w', encoding='utf-8') as txt_file_content:
            for line in lines:
                # Ekstrak nama
                name_match = re.search(r'FN:(.+)', line)
                # Ekstrak nomor telepon
                tel_match = re.search(r'TEL;TYPE=CELL:(\+?\d+)', line)

                if tel_match:
                    tel = tel_match.group(1).strip()
                    tel = re.sub(r'\D', '', tel)  # Menghapus semua karakter non-digit

                    # Menulis nama dan nomor telepon ke dalam file teks
                    txt_file_content.write(f"{tel}\n")

        return txt_file
    except Exception as e:
        logging.error("Error converting VCF to TXT: ", exc_info=True)
        raise


def gabung_vcf(input_files, output_file):
    logging.info("Memulai penggabungan VCF.")
    with open(output_file, 'wb') as outfile:
        for i, filename in enumerate(input_files):
            logging.info(f"Membaca file: {filename}")
            with open(filename, 'rb') as infile:
                content = infile.read()
                outfile.write(content)
                
            if i < len(input_files) - 1:
                outfile.write(b'\n')

    logging.info(f"Penggabungan selesai. File output: {output_file}")
    

# Fungsi untuk menyimpan teks ke file TXT
def save_txt(text, filename):
    file_path = os.path.join('files', filename)
    with open(file_path, 'w') as file:
        file.write(text)
    return file_path

# Fungsi untuk menjalankan perintah shell dan mengembalikan output
def run_command(command):
    try:
        output = subprocess.check_output(command, shell=True, stderr=subprocess.STDOUT)
        return output.decode('utf-8')
    except subprocess.CalledProcessError as e:
        logging.error(f"Command failed: {e}")
        return None

# Fungsi untuk mengeksploitasi WiFi WPS/WPA menggunakan alat `reaver`
def exploit_wifi_wps(interface, bssid, channel):
    try:
        # Set interface ke mode monitor
        run_command(f"airmon-ng start {interface} {channel}")
        
        # Jalankan reaver untuk eksploitasi WPS
        reaver_command = f"reaver -i {interface} -b {bssid} -c {channel} -vv"
        reaver_output = run_command(reaver_command)
        
        # Ekstrak informasi dari output reaver
        ssid = extract_ssid(reaver_output)
        pin = extract_pin(reaver_output)
        password = extract_password(reaver_output)
        security = "WPA/WPA2"  # Logika bisa ditambahkan untuk mendeteksi jenis keamanan lain

        # Asumsikan kelemahan berdasarkan apakah WPS aktif
        weakness = "WPS Enabled" if "WPS" in reaver_output else "No WPS"

        return {
            'ssid': ssid or "Unknown",
            'pin': pin or "Not Found",
            'password': password or "Not Found",
            'security': security,
            'weakness': weakness
        }
    except Exception as e:
        logging.error(f"Error in exploit_wifi_wps: {e}")
        return {
            'ssid': "Unknown",
            'pin': "Not Found",
            'password': "Not Found",
            'security': "Unknown",
            'weakness': "Unknown"
        }

# Fungsi untuk mengekstrak SSID dari output reaver
def extract_ssid(output):
    ssid_match = re.search(r"SSID\s+:\s+(.+)", output)
    if ssid_match:
        return ssid_match.group(1).strip()
    return None

# Fungsi untuk mengekstrak PIN dari output reaver
def extract_pin(output):
    pin_match = re.search(r"WPS PIN:\s+(\d{8})", output)
    if pin_match:
        return pin_match.group(1)
    return None

# Fungsi untuk mengekstrak password dari output reaver
def extract_password(output):
    password_match = re.search(r"PSK\s+:\s+(.+)", output)
    if password_match:
        return password_match.group(1).strip()
    return None

def gabung_txt(input_files, output_file):
    logging.info("Memulai penggabungan TXT.")
    with open(output_file, 'w', encoding='utf-8') as outfile:
        for i, filename in enumerate(input_files):
            logging.info(f"Membaca file: {filename}")
            with open(filename, 'r', encoding='utf-8') as infile:
                content = infile.read()
                outfile.write(content)
                
            if i < len(input_files) - 1:
                outfile.write('\n')  # Tambahkan baris baru antara file jika bukan file terakhir

    logging.info(f"Penggabungan selesai. File output: {output_file}")

def remove_plus_and_spaces(input_file, output_file):
    logging.info(f"Memproses file: {input_file}")
    with open(input_file, 'r') as infile, open(output_file, 'w') as outfile:
        for line in infile:
            # Hapus semua tanda '+' dan spasi
            cleaned_line = line.replace('+', '').replace(' ', '')
            outfile.write(cleaned_line)
    logging.info(f"File selesai diproses: {output_file}")

def gabungkan_kolom(input_file, output_file):
    """
    Fungsi untuk menggabungkan kolom-kolom dari atas ke kiri lalu kebawah,
    lalu ke kolom selanjutnya dan seterusnya sampai isi file habis.
    """
    try:
        with open(input_file, 'r', newline='', encoding='utf-8') as f_input:
            reader = csv.reader(f_input, delimiter='\t')  # Anggap file menggunakan delimiter tab
            data = list(reader)
        
        # Menyiapkan list untuk menampung hasil penggabungan
        combined_col = []
        
        # Dapatkan jumlah baris dan kolom
        max_rows = len(data)
        max_cols = len(data[0]) if max_rows > 0 else 0

        # Proses penggabungan dari atas ke kiri, lalu kebawah dan ke kolom berikutnya
        for row_index in range(max_rows):
            for col_index in range(max_cols):
                for r in range(row_index, max_rows):
                    if col_index < len(data[r]):  # Pastikan kolom ada di baris ini
                        combined_col.append(data[r][col_index])

        # Tulis hasilnya ke file output
        with open(output_file, 'w', newline='', encoding='utf-8') as f_output:
            writer = csv.writer(f_output, delimiter='\t')
            for item in combined_col:
                writer.writerow([item])  # Setiap elemen ditulis di baris baru

    except Exception as e:
        print(f"Error saat menggabungkan kolom: {e}")


def split(arr, num):
    return [arr[x:x+num] for x in range(0, len(arr), num)]

if __name__ == "__main__":
    data = {
        'filename': 'files/11-112.txt',
        'name': 'tes',
        'cname': 'tes',
        'totalc': 100,
        'totalf': 5,
    }
    convert(data)
